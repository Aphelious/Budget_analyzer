from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, datetime
from env import *
import json
from datetime import date
import dateparser
# import nltk
# from nltk import word_tokenize, sent_tokenize
# from pprint import pprint




wb = load_workbook(filename= filename)
wb.iso_dates = True
ws = wb.active



def create_new_cols():

    '''Create 4 new columns on the left side of the spreadsheet (cols A - D). Column A is purely a cosmetic
    buffer on the left side of the sheet (width 4). Columns B - D are meant for extra date data (width 12). Also
    delete the header row.'''

    ws.insert_cols(0, 4)
    ws.delete_rows(1)



def format_col_widths():
    '''Set column widths to accept new data and display long strings in Column H (width 80).'''

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['H'].width = 80



def set_alignment(cols):

    '''Set alignment of columns passed in to 'center'.

    ::param:: cols - tuple of strings representing column names in destination spreadsheet

        ex: ('B', 'C', 'D')
    '''

    wb2 = load_workbook(filename=master_file)
    ws2 = wb2["Raw Data"]
    dest_max_column = ws2.max_column
    dest_max_row = ws2.max_row
    for row in range(1, dest_max_row + 1):
        for i in range(1, dest_max_column + 1):
            col = get_column_letter(i)
            ws2[f'{col}{row}'].alignment = Alignment(horizontal='center')
    wb2.save(filename=master_file)
    wb2.close()



def format_dates():

    '''Format date data and insert into newly-created columns. '''

    month_hash = {
        1: 'Jan',
        2: 'Feb',
        3: 'March',
        4: 'April',
        5: 'May',
        6: 'June',
        7: 'July',
        8: 'Aug',
        9: 'Sept',
        10: 'Oct',
        11: 'Nov',
        12: 'Dec'}

    for row in range(1, 1000):
        if ws[f'E{row}'].value == None:
            break
        date = ws[f'E{row}'].value
        ws[f'B{row}'].value = date.year
        ws[f'C{row}'].value = f'{date.year} - {month_hash[date.month]}'
        ws[f'D{row}'].value = 'Long Island'
        ws[f'E{row}'].value = date.date()



def combine_debits_credits():

    '''Combine Debit and Credit columns into a single column and preserve sign of values (+ or -).'''

    for row in range(1, 1000):
        if ws[f'E{row}'].value == None:
            break
        amount = ws[f'J{row}'].value
        if amount and ws[f'I{row}'].value == None:
            ws[f'I{row}'].value = amount



def delete_unused_cols():

    '''Delete column J after 'combine_debits_credits()' has been run to move credits values over to
    column I.'''

    ws.delete_cols(10)



def split_cell_values(column):

    '''Split all distinct alpha-numeric sequences in Column H (Description), append the resulting list to a list
    of lists, called 'text', and return it.'''

    text = []  # each line in the spreadsheet is stored here in separate lists
    for cell in ws['H']:
        words = [x.lower() for x in cell.value.split() if not x.isdigit()]  # Only keep continuous string sequences if
        text.append(words)  # they are not entirely numbers
    return text



def words_hash(text):

    '''Creates hash entry (if one doesn't exist) that records the count of each textual sequence used in the column of the
    spreadsheet and returns a list, sorted by most common word.'''

    words_hash = {}  # each word in the collective description column will be counted and hashed here
    for line in text:
        for word in line:
            if word in words_hash:
                words_hash[word] += 1
            else:
                words_hash.update({word: 1})
    sorted_dict = (sorted(words_hash, key=words_hash.get, reverse=True))  # Sort the dictionary by the frequency of usage
    return sorted_dict



def categorize_description(sorted_dict, text):

    '''Handles the mapping of user-defined categories to unique textual sequences by printing out the lines in
    which the sequence occurs and then prompting the user for input. Results are stored in a database 'descriptions.db'
    or a json file 'descriptions.json' for later recall.'''

    try:
        with open(descriptions_filepath) as f:
            descript_hash = json.load(f)
            for word in sorted_dict:
                if word not in descript_hash:
                    print(f'Here are the lines in which \'{word}\' occurs in the spreadsheet.')
                    for line in text:
                        if word in line:
                            print(line)
                    description = input(f'What description do you want to associate with the term \'{word}\' ?')
                    descript_hash.update({word: description})
            with open(descriptions_filepath, 'w') as f:
                json.dump(descript_hash, f)
            return descript_hash
    except:
        print('No previous description file found. Creating new...')

    descript_hash = {}  # Store user-defined categories here
    for word in sorted_dict:
        if word not in descript_hash:
            print(f'Here are the lines in which \'{word}\' occurs in the spreadsheet.')
            for line in text:
                if word in line:
                    print(line)
            description = input(f'What description do you want to associate with the term \'{word}\' ?')
            descript_hash.update({word: description})
    breakpoint()
    with open(descriptions_filepath, 'w') as f:
        json.dump(descript_hash, f)
    return descript_hash



def write_description(descript_hash, text):
    line_ref = 0
    for line in text:
        line_ref += 1
        for word in line:
            if word in descript_hash and descript_hash[word] != 'None':
                print(f'Description found for word \'{word}\': \'{descript_hash[word]}\'. Writing value to spreadsheet.')
                ws[f'F{line_ref}'].value = descript_hash[word]
            else:
                continue



def check_description_edge_cases():
    line_ref = 0
    for cell in ws['F']:
        line_ref += 1
        column_text = ws[f'H{line_ref}'].value
        amount = ws[f'I{line_ref}'].value
        if cell.value == None:
            description = input(f'What description do you want for this edge case I found: '
                                f'\n\'{column_text}\' for ${amount}?')
            print(f'Okay, writing \'{description}\' to file.')
            ws[f'F{line_ref}'].value = description
        else:
            continue


def categorize_meta_description():

    '''Group some categories into meta-categories such as 'Food' representing all food-related categories. '''

    line_ref = 0
    food = ['Groceries', 'Fast food', 'Pizza', 'Deli', 'Seamless', 'Grubhub', 'Chinese food', 'food', 'Food',
            'Out to eat', 'Breakfast', 'Dessert', 'Diner', 'Factor 75', 'Freshly', 'Lunch', 'Taco Bell',
            'Vending machine', 'Yes Please']

    car_costs = ['Gas', 'gas', 'Car insurance', 'Car Insurance', 'Car repairs', 'Car Repairs', 'Car parts',
                 'Car Document Fees', 'Car maintenance']
    for cell in ws['F']:
        line_ref += 1
        if cell.value in food:
            ws[f'G{line_ref}'].value = 'Food'
        elif cell.value in car_costs:
            ws[f'G{line_ref}'].value = 'Car costs'
        else:
            ws[f'G{line_ref}'].value = None



def set_sheet_style():
    '''Set destination spreadsheet font to 'Calibri' and font size 14.'''

    font = Font(name='Calibri',
                size=14)
    wb2 = load_workbook(filename=master_file)
    wb2.iso_dates = True
    ws2 = wb2["Raw Data"]
    dest_max_column = ws2.max_column
    dest_max_row = ws2.max_row
    for row in range(1, dest_max_row + 1):
        for i in range(1, dest_max_column + 1):
            col = get_column_letter(i)
            ws2[f'{col}{row}'].font = font

    wb2.save(filename=master_file)
    wb2.close()


def set_iso_dates():
    '''Set dates in column 'E' to ISO format. Reads values into openpyxl and writes the string outputs of
    those datetime objects.
    '''

    wb2 = load_workbook(filename=master_file)
    wb2.iso_dates = True
    ws2 = wb2["Raw Data"]
    for cell in ws2['E'][1:]:
        cell_date = str(cell.value)
        cell.value = cell_date

    wb2.save(filename=master_file)
    wb2.close()


def parse_description(column):

    text = split_cell_values(column)
    sorted_dict = words_hash(text)
    descript_hash = categorize_description(sorted_dict, text)
    write_description(descript_hash, text)
    check_description_edge_cases()
    categorize_meta_description()



def append_to_master_xslx():

    '''Write formatted and categorized lines to master spreadsheet'''

    wb2 = load_workbook(filename=master_file)
    ws2 = wb2["Raw Data"]
    wb2.iso_dates = True
    dest_max_row = ws2.max_row
    source_min_column = ws.min_column
    source_max_column = ws.max_column
    source_min_row = ws.min_row
    source_max_row = ws.max_row
    for row in range(source_min_row, source_max_row + 1):
        for i in range(source_min_column, source_max_column + 1):
            col = get_column_letter(i)
            value = ws[f'{col}{row}'].value
            if type(value) is datetime.datetime:
                value = value.date()
            ws2[f'{col}{dest_max_row + row}'].value = value

    wb2.save(filename=master_file)
    wb2.close()



def format_spreadsheet():
    create_new_cols()
    format_col_widths()
    format_dates()
    combine_debits_credits()
    delete_unused_cols()
    parse_description('H')
    set_sheet_style()



def format_dest_spreadsheet():

    set_alignment(('B', 'C', 'D', 'F', 'G'))
    set_iso_dates()
    set_sheet_style()



#############
# Dashboard #
#############


# create_new_cols()
# format_col_widths()
# set_alignment(('B', 'C', 'D'))
# format_dates()
# combine_debits_credits()
# delete_unused_cols()
# parse_description('H')
# set_sheet_style()
# set_alignment(('B', 'C', 'D', 'F', 'G'))
# set_iso_dates()

# format_spreadsheet()
# append_to_master_xslx()
# format_dest_spreadsheet()



wb.save(filename=filename)
wb.close()